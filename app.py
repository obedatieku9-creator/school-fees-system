from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from flask_wtf import CSRFProtect
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import generate_password_hash
import os
from models.database import Database, User, Student, Class, FeeStructure, Payment, Report
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
import openpyxl
from datetime import datetime

class Config:
    SECRET_KEY = os.environ.get('SECRET_KEY') or 'replace-this-secret-key'
    SESSION_COOKIE_HTTPONLY = True
    REMEMBER_COOKIE_HTTPONLY = True
    SESSION_COOKIE_SAMESITE = 'Lax'
    PREFERRED_URL_SCHEME = 'https'
    DEBUG = False
    TESTING = False

class ProductionConfig(Config):
    SESSION_COOKIE_SECURE = True
    REMEMBER_COOKIE_SECURE = True

class DevelopmentConfig(Config):
    DEBUG = True
    SESSION_COOKIE_SECURE = False
    REMEMBER_COOKIE_SECURE = False

app = Flask(__name__)
env = os.environ.get('FLASK_ENV', 'development').lower()
app.config.from_object(DevelopmentConfig if env == 'development' else ProductionConfig)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1)
csrf = CSRFProtect(app)
app.csrf = csrf

if app.config['SECRET_KEY'] == 'replace-this-secret-key':
    app.logger.warning('Using default SECRET_KEY. Set SECRET_KEY in environment for production.')

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

class AdminUser(UserMixin):
    def __init__(self, id):
        self.id = id

@login_manager.user_loader
def load_user(user_id):
    return AdminUser(user_id)

@app.context_processor
def utility_processor():
    def get_item(obj, key, default=''):
        try:
            return obj[key]
        except (KeyError, IndexError, TypeError):
            return default
    return dict(get_item=get_item)

@app.route('/')
@login_required
def dashboard():
    stats = Report.get_dashboard_stats()
    return render_template('dashboard.html', stats=stats)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if User.authenticate(username, password):
            user = AdminUser(username)
            login_user(user)
            return redirect(url_for('dashboard'))
        flash('Invalid credentials')
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current_password = request.form['current_password']
        new_password = request.form['new_password']
        confirm_password = request.form['confirm_password']

        if new_password != confirm_password:
            flash('New passwords do not match')
            return redirect(url_for('change_password'))

        if not User.authenticate(current_user.id, current_password):
            flash('Current password is incorrect')
            return redirect(url_for('change_password'))

        User.set_password(current_user.id, new_password)
        flash('Password updated successfully')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')

@app.route('/students')
@login_required
def students():
    students = Student.get_all()
    return render_template('students.html', students=students)

@app.route('/students/add', methods=['GET', 'POST'])
@login_required
def add_student():
    classes = Class.get_all()
    if request.method == 'POST':
        student_id = Student.generate_student_id()
        data = (
            student_id,
            request.form['full_name'],
            request.form['gender'],
            request.form['date_of_birth'],
            request.form['class_id'],
            request.form['parent_name'],
            request.form['parent_phone'],
            request.form['address']
        )
        Student.add(data)
        flash('Student added successfully')
        return redirect(url_for('students'))
    return render_template('add_student.html', classes=classes)

@app.route('/students/edit/<student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    student = Student.get_by_id(student_id)
    classes = Class.get_all()
    if request.method == 'POST':
        data = (
            request.form['full_name'],
            request.form['gender'],
            request.form['date_of_birth'],
            request.form['class_id'],
            request.form['parent_name'],
            request.form['parent_phone'],
            request.form['address']
        )
        Student.update(student_id, data)
        flash('Student updated successfully')
        return redirect(url_for('students'))
    return render_template('edit_student.html', student=student, classes=classes)

@app.route('/students/delete/<student_id>')
@login_required
def delete_student(student_id):
    Student.delete(student_id)
    flash('Student deleted successfully')
    return redirect(url_for('students'))

@app.route('/students/search')
@login_required
def search_students():
    query = request.args.get('q', '')
    students = Student.search(query)
    return render_template('students.html', students=students, search_query=query)

@app.route('/students/import', methods=['POST'])
@login_required
def import_students():
    if 'file' not in request.files:
        flash('No file selected')
        return redirect(url_for('students'))
    
    file = request.files['file']
    if file.filename == '':
        flash('No file selected')
        return redirect(url_for('students'))
    
    if file and (file.filename.endswith('.csv') or file.filename.endswith(('.xlsx', '.xls'))):
        try:
            imported_count = Student.import_from_file(file)
            flash(f'Successfully imported {imported_count} students')
        except Exception as e:
            flash(f'Import failed: {str(e)}')
    else:
        flash('Invalid file type. Please upload CSV or Excel file.')
    
    return redirect(url_for('students'))

@app.route('/students/export')
@login_required
def export_students():
    students = Student.get_all()
    buffer = Student.export_to_excel(students)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=students.xlsx'
    return response

@app.route('/classes')
@login_required
def classes():
    classes = Class.get_all()
    return render_template('classes.html', classes=classes)

@app.route('/fees')
@login_required
def fees():
    fees = FeeStructure.get_all()
    classes = Class.get_all()
    return render_template('fees.html', fees=fees, classes=classes)

@app.route('/fees/set', methods=['POST'])
@login_required
def set_fees():
    class_id = request.form['class_id']
    term = request.form['term']
    amount = float(request.form['amount'])
    FeeStructure.set_fee(class_id, term, amount)
    flash('Fee structure updated successfully')
    return redirect(url_for('fees'))

@app.route('/payments')
@login_required
def payments():
    payments = Payment.get_all()
    return render_template('payments.html', payments=payments)

@app.route('/payments/add', methods=['GET', 'POST'])
@login_required
def add_payment():
    students = Student.get_all()
    if request.method == 'POST':
        receipt_number = Payment.generate_receipt_number()
        data = (
            request.form['student_id'],
            float(request.form['amount']),
            request.form['term'],
            request.form['payment_date'],
            request.form['payment_method'],
            receipt_number
        )
        Payment.add(data)
        flash(f'Payment recorded successfully. Receipt: {receipt_number}')
        return redirect(url_for('receipt', receipt_number=receipt_number))
    return render_template('add_payment.html', students=students)

@app.route('/receipt/<receipt_number>')
@login_required
def receipt(receipt_number):
    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(db.convert_query('''
        SELECT p.*, s.full_name, s.student_id, c.name as class_name
        FROM payments p
        JOIN students s ON p.student_id = s.student_id
        JOIN classes c ON s.class_id = c.id
        WHERE p.receipt_number = ?
    '''), (receipt_number,))
    payment = cursor.fetchone()
    conn.close()

    if payment:
        balance = Payment.get_student_balance(payment[1])  # student_id
        return render_template('receipt.html', payment=payment, balance=balance)
    return 'Receipt not found', 404

@app.route('/reports')
@login_required
def reports():
    return render_template('reports.html')

@app.route('/reports/fees_per_class')
@login_required
def fees_per_class():
    data = Report.get_fees_per_class()
    return render_template('report_fees_per_class.html', data=data)

@app.route('/reports/students_owing')
@login_required
def students_owing():
    data = Report.get_students_owing()
    return render_template('report_students_owing.html', data=data)

@app.route('/reports/fees_per_term')
@login_required
def fees_per_term():
    data = Report.get_fees_per_term()
    return render_template('report_fees_per_term.html', data=data)

@app.route('/export/<report_type>/<format>')
@login_required
def export_report(report_type, format):
    if report_type == 'fees_per_class':
        data = Report.get_fees_per_class()
        headers = ['Class', 'Total Fees Collected']
    elif report_type == 'students_owing':
        data = Report.get_students_owing()
        headers = ['Student ID', 'Name', 'Class', 'Balance']
    elif report_type == 'fees_per_term':
        data = Report.get_fees_per_term()
        headers = ['Term', 'Total Fees Collected']

    if format == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        styles = getSampleStyleSheet()
        elements.append(Paragraph(f"School Fees Report - {report_type.replace('_', ' ').title()}", styles['Title']))
        elements.append(Spacer(1, 12))

        table_data = [headers] + list(data)
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename={report_type}_report.pdf'
        return response

    elif format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_type.replace('_', ' ').title()

        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        for row_num, row in enumerate(data, 2):
            for col_num, cell_value in enumerate(row, 1):
                ws.cell(row=row_num, column=col_num, value=cell_value)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={report_type}_report.xlsx'
        return response

@app.route('/analytics')
@login_required
def analytics():
    # Get analytics data
    collection_trends = Report.get_collection_trends()
    payment_methods = Report.get_payment_methods_distribution()
    class_performance = Report.get_class_performance()
    outstanding_fees = Report.get_outstanding_fees_by_class()
    monthly_patterns = Report.get_monthly_payment_patterns()
    
    return render_template('analytics.html',
                         collection_trends=collection_trends,
                         payment_methods=payment_methods,
                         class_performance=class_performance,
                         outstanding_fees=outstanding_fees,
                         monthly_patterns=monthly_patterns)

if __name__ == '__main__':
    # Initialize admin user if not exists
    db = Database()
    db.init_admin()
    app.run(debug=True)