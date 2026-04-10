import os
import shutil
import sqlite3
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime

try:
    import psycopg2  # type: ignore[import]
    import psycopg2.extras  # type: ignore[import]
except ImportError:
    psycopg2 = None


class DBRow(tuple):
    def __new__(cls, values, keys):
        obj = super().__new__(cls, values)
        obj._keys = list(keys)
        obj._key_index = {key: index for index, key in enumerate(keys)}
        return obj

    def __getitem__(self, key):
        if isinstance(key, str):
            index = self._key_index.get(key)
            if index is not None:
                return super().__getitem__(index)
        return super().__getitem__(key)

    def get(self, key, default=None):
        try:
            return self[key]
        except Exception:
            return default


class Database:
    def __init__(self):
        default_db = os.path.join(os.path.dirname(__file__), '..', 'database', 'school_fees.db')
        self.db_url = os.environ.get('DATABASE_URL')
        self.use_postgres = bool(self.db_url)
        self.db_path = os.environ.get('DATABASE_PATH', default_db)

        if not self.use_postgres:
            db_dir = os.path.dirname(self.db_path)
            if db_dir and not os.path.exists(db_dir):
                os.makedirs(db_dir, exist_ok=True)
            if self.db_path != default_db and not os.path.exists(self.db_path):
                if os.path.exists(default_db):
                    shutil.copyfile(default_db, self.db_path)

        self.ensure_database()

    def get_connection(self):
        if self.use_postgres:
            if psycopg2 is None:
                raise RuntimeError('psycopg2 is required for PostgreSQL support')
            return psycopg2.connect(self.db_url)

        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def convert_query(self, query):
        if self.use_postgres:
            return query.replace('?', '%s')
        return query

    def _wrap_one(self, cursor, row):
        if row is None or not self.use_postgres:
            return row
        keys = [d[0] for d in cursor.description]
        return DBRow(row, keys)

    def _wrap_all(self, cursor, rows):
        if not self.use_postgres:
            return rows
        keys = [d[0] for d in cursor.description]
        return [DBRow(row, keys) for row in rows]

    def has_table(self, table_name):
        conn = self.get_connection()
        cursor = conn.cursor()
        if self.use_postgres:
            cursor.execute(
                "SELECT EXISTS (SELECT 1 FROM information_schema.tables WHERE table_schema='public' AND table_name=%s)",
                (table_name,)
            )
            exists = cursor.fetchone()[0]
        else:
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
            exists = cursor.fetchone() is not None
        conn.close()
        return exists

    def initialize_database(self):
        conn = self.get_connection()
        cursor = conn.cursor()

        if self.use_postgres:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id SERIAL PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT DEFAULT 'admin'
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS classes (
                    id SERIAL PRIMARY KEY,
                    name TEXT UNIQUE NOT NULL,
                    section TEXT NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id SERIAL PRIMARY KEY,
                    student_id TEXT UNIQUE NOT NULL,
                    full_name TEXT NOT NULL,
                    gender TEXT NOT NULL,
                    date_of_birth DATE NOT NULL,
                    class_id INTEGER NOT NULL REFERENCES classes(id),
                    parent_name TEXT NOT NULL,
                    parent_phone TEXT NOT NULL,
                    address TEXT NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS fees_structure (
                    id SERIAL PRIMARY KEY,
                    class_id INTEGER NOT NULL REFERENCES classes(id),
                    term TEXT NOT NULL,
                    amount REAL NOT NULL,
                    UNIQUE(class_id, term)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS payments (
                    id SERIAL PRIMARY KEY,
                    student_id TEXT NOT NULL,
                    amount REAL NOT NULL,
                    term TEXT NOT NULL,
                    payment_date DATE NOT NULL,
                    payment_method TEXT NOT NULL,
                    receipt_number TEXT UNIQUE NOT NULL,
                    FOREIGN KEY (student_id) REFERENCES students (student_id)
                )
            ''')

            cursor.execute(
                '''
                INSERT INTO users (username, password_hash, role)
                VALUES (%s, %s, %s)
                ON CONFLICT (username) DO NOTHING
                ''',
                ('admin', generate_password_hash('admin123'), 'admin')
            )

            classes = [
                ('Foundation 1', 'Pre-School'),
                ('Foundation 2', 'Pre-School'),
                ('Reception 1', 'Pre-School'),
                ('Reception 2', 'Pre-School'),
                ('Primary 1', 'Primary School'),
                ('Primary 2', 'Primary School'),
                ('Primary 3', 'Primary School'),
                ('Primary 4', 'Primary School'),
                ('Primary 5', 'Primary School'),
                ('Primary 6', 'Primary School'),
                ('JHS 1', 'JHS'),
                ('JHS 2', 'JHS'),
                ('JHS 3', 'JHS')
            ]

            for name, section in classes:
                cursor.execute(
                    '''
                    INSERT INTO classes (name, section)
                    VALUES (%s, %s)
                    ON CONFLICT (name) DO NOTHING
                    ''',
                    (name, section)
                )
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT DEFAULT 'admin'
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS classes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT UNIQUE NOT NULL,
                    section TEXT NOT NULL
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS students (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id TEXT UNIQUE NOT NULL,
                    full_name TEXT NOT NULL,
                    gender TEXT NOT NULL,
                    date_of_birth TEXT NOT NULL,
                    class_id INTEGER NOT NULL,
                    parent_name TEXT NOT NULL,
                    parent_phone TEXT NOT NULL,
                    address TEXT NOT NULL,
                    FOREIGN KEY (class_id) REFERENCES classes (id)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS fees_structure (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    class_id INTEGER NOT NULL,
                    term TEXT NOT NULL,
                    amount REAL NOT NULL,
                    FOREIGN KEY (class_id) REFERENCES classes (id),
                    UNIQUE(class_id, term)
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS payments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    student_id TEXT NOT NULL,
                    amount REAL NOT NULL,
                    term TEXT NOT NULL,
                    payment_date TEXT NOT NULL,
                    payment_method TEXT NOT NULL,
                    receipt_number TEXT UNIQUE NOT NULL,
                    FOREIGN KEY (student_id) REFERENCES students (student_id)
                )
            ''')

            cursor.execute('''
                INSERT OR IGNORE INTO users (username, password_hash, role)
                VALUES (?, ?, ?)
            ''', ('admin', generate_password_hash('admin123'), 'admin'))

            classes = [
                ('Foundation 1', 'Pre-School'),
                ('Foundation 2', 'Pre-School'),
                ('Reception 1', 'Pre-School'),
                ('Reception 2', 'Pre-School'),
                ('Primary 1', 'Primary School'),
                ('Primary 2', 'Primary School'),
                ('Primary 3', 'Primary School'),
                ('Primary 4', 'Primary School'),
                ('Primary 5', 'Primary School'),
                ('Primary 6', 'Primary School'),
                ('JHS 1', 'JHS'),
                ('JHS 2', 'JHS'),
                ('JHS 3', 'JHS')
            ]

            for name, section in classes:
                cursor.execute('INSERT OR IGNORE INTO classes (name, section) VALUES (?, ?)', (name, section))

        conn.commit()
        conn.close()

    def ensure_database(self):
        if self.use_postgres:
            if not self.has_table('users'):
                self.initialize_database()
            return

        if not os.path.exists(self.db_path):
            self.initialize_database()
            return

        if not self.has_table('users'):
            self.initialize_database()

    def init_admin(self, password='admin123'):
        conn = self.get_connection()
        cursor = conn.cursor()
        hashed = generate_password_hash(password)
        cursor.execute(self.convert_query('UPDATE users SET password_hash = ? WHERE username = ?'), (hashed, 'admin'))
        conn.commit()
        conn.close()


class User:
    @staticmethod
    def authenticate(username, password):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('SELECT password_hash FROM users WHERE username = ?'), (username,))
        result = db._wrap_one(cursor, cursor.fetchone())
        conn.close()
        if result and check_password_hash(result[0], password):
            return True
        return False

    @staticmethod
    def set_password(username, new_password):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        hashed = generate_password_hash(new_password)
        cursor.execute(db.convert_query('UPDATE users SET password_hash = ? WHERE username = ?'), (hashed, username))
        conn.commit()
        conn.close()

class Student:
    @staticmethod
    def get_all():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT s.*, c.name as class_name
            FROM students s
            JOIN classes c ON s.class_id = c.id
            ORDER BY s.full_name
        ''')
        students = db._wrap_all(cursor, cursor.fetchall())
        conn.close()
        return students

    @staticmethod
    def get_by_id(student_id):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('''
            SELECT s.*, c.name as class_name
            FROM students s
            JOIN classes c ON s.class_id = c.id
            WHERE s.student_id = ?
        '''), (student_id,))
        student = db._wrap_one(cursor, cursor.fetchone())
        conn.close()
        return student

    @staticmethod
    def search(query):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('''
            SELECT s.*, c.name as class_name
            FROM students s
            JOIN classes c ON s.class_id = c.id
            WHERE s.full_name LIKE ? OR s.student_id LIKE ? OR c.name LIKE ?
        '''), (f'%{query}%', f'%{query}%', f'%{query}%'))
        students = db._wrap_all(cursor, cursor.fetchall())
        conn.close()
        return students

    @staticmethod
    def add(student_data):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('''
            INSERT INTO students (student_id, full_name, gender, date_of_birth, class_id, parent_name, parent_phone, address)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        '''), student_data)
        conn.commit()
        conn.close()

    @staticmethod
    def update(student_id, student_data):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('''
            UPDATE students
            SET full_name = ?, gender = ?, date_of_birth = ?, class_id = ?, parent_name = ?, parent_phone = ?, address = ?
            WHERE student_id = ?
        '''), student_data + (student_id,))
        conn.commit()
        conn.close()

    @staticmethod
    def delete(student_id):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('DELETE FROM students WHERE student_id = ?'), (student_id,))
        conn.commit()
        conn.close()

    @staticmethod
    def generate_student_id():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM students')
        count = cursor.fetchone()[0]
        conn.close()
        return f'ST{str(count + 1).zfill(4)}'

    @staticmethod
    def import_from_file(file):
        import csv
        import io
        
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        # Get class mapping
        cursor.execute('SELECT id, name FROM classes')
        class_map = {row[1]: row[0] for row in db._wrap_all(cursor, cursor.fetchall())}
        
        imported_count = 0
        
        if file.filename.endswith('.csv'):
            # CSV handling
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
        else:
            # Excel handling
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            reader = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                reader.append(dict(zip(headers, row)))
        
        for row in reader:
            try:
                class_name = row.get('class_name', '').strip()
                if class_name not in class_map:
                    continue  # Skip if class not found
                
                student_id = Student.generate_student_id()
                student_data = (
                    student_id,
                    row.get('full_name', '').strip(),
                    row.get('gender', '').strip(),
                    row.get('date_of_birth', '').strip(),
                    class_map[class_name],
                    row.get('parent_name', '').strip(),
                    row.get('parent_phone', '').strip(),
                    row.get('address', '').strip()
                )
                
                cursor.execute(db.convert_query('''
                    INSERT INTO students (student_id, full_name, gender, date_of_birth, class_id, parent_name, parent_phone, address)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                '''), student_data)
                imported_count += 1
                
            except Exception:
                continue  # Skip invalid rows
        
        conn.commit()
        conn.close()
        return imported_count

    @staticmethod
    def export_to_excel(students):
        from openpyxl import Workbook
        from io import BytesIO
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        
        # Headers
        headers = ['Student ID', 'Full Name', 'Gender', 'Date of Birth', 'Class', 'Parent Name', 'Parent Phone', 'Address']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Data
        for row_num, student in enumerate(students, 2):
            ws.cell(row=row_num, column=1, value=student[1])  # student_id
            ws.cell(row=row_num, column=2, value=student[2])  # full_name
            ws.cell(row=row_num, column=3, value=student[3])  # gender
            ws.cell(row=row_num, column=4, value=student[4])  # date_of_birth
            ws.cell(row=row_num, column=5, value=student[9])  # class_name
            ws.cell(row=row_num, column=6, value=student[6])  # parent_name
            ws.cell(row=row_num, column=7, value=student[7])  # parent_phone
            ws.cell(row=row_num, column=8, value=student[8])  # address
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

class Class:
    @staticmethod
    def get_all():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM classes ORDER BY name')
        classes = db._wrap_all(cursor, cursor.fetchall())
        conn.close()
        return classes

class FeeStructure:
    @staticmethod
    def get_all():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT fs.*, c.name as class_name
            FROM fees_structure fs
            JOIN classes c ON fs.class_id = c.id
            ORDER BY c.name, fs.term
        ''')
        fees = db._wrap_all(cursor, cursor.fetchall())
        conn.close()
        return fees

    @staticmethod
    def set_fee(class_id, term, amount):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        if db.use_postgres:
            cursor.execute('''
                INSERT INTO fees_structure (class_id, term, amount)
                VALUES (%s, %s, %s)
                ON CONFLICT (class_id, term) DO UPDATE SET amount = EXCLUDED.amount
            ''', (class_id, term, amount))
        else:
            cursor.execute('''
                INSERT OR REPLACE INTO fees_structure (class_id, term, amount)
                VALUES (?, ?, ?)
            ''', (class_id, term, amount))
        conn.commit()
        conn.close()

class Payment:
    @staticmethod
    def get_all():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT p.*, s.full_name, c.name as class_name
            FROM payments p
            JOIN students s ON p.student_id = s.student_id
            JOIN classes c ON s.class_id = c.id
            ORDER BY p.payment_date DESC
        ''')
        payments = db._wrap_all(cursor, cursor.fetchall())
        conn.close()
        return payments

    @staticmethod
    def get_by_student(student_id):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('''
            SELECT * FROM payments WHERE student_id = ? ORDER BY payment_date DESC
        '''), (student_id,))
        payments = db._wrap_all(cursor, cursor.fetchall())
        conn.close()
        return payments

    @staticmethod
    def add(payment_data):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute(db.convert_query('''
            INSERT INTO payments (student_id, amount, term, payment_date, payment_method, receipt_number)
            VALUES (?, ?, ?, ?, ?, ?)
        '''), payment_data)
        conn.commit()
        conn.close()

    @staticmethod
    def generate_receipt_number():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM payments')
        count = cursor.fetchone()[0]
        conn.close()
        return f'RCP{str(count + 1).zfill(6)}'

    @staticmethod
    def get_student_balance(student_id):
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()

        # Get total fees for student's class and terms
        cursor.execute(db.convert_query('''
            SELECT fs.term, fs.amount
            FROM fees_structure fs
            JOIN students s ON fs.class_id = s.class_id
            WHERE s.student_id = ?
        '''), (student_id,))
        fees = cursor.fetchall()

        # Get total payments
        cursor.execute(db.convert_query('''
            SELECT term, SUM(amount) as total_paid
            FROM payments
            WHERE student_id = ?
            GROUP BY term
        '''), (student_id,))
        payments = cursor.fetchall()

        conn.close()

        balance = {}
        for term, amount in fees:
            paid = sum(p[1] for p in payments if p[0] == term)
            balance[term] = amount - paid

        return balance

class Report:
    @staticmethod
    def get_dashboard_stats():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()

        # Total students
        cursor.execute('SELECT COUNT(*) FROM students')
        total_students = cursor.fetchone()[0]

        # Total fees collected
        cursor.execute('SELECT SUM(amount) FROM payments')
        total_collected = cursor.fetchone()[0] or 0

        # Outstanding balances
        cursor.execute('''
            SELECT SUM(fs.amount - COALESCE(p.total_paid, 0)) as outstanding
            FROM fees_structure fs
            LEFT JOIN (
                SELECT class_id, term, SUM(amount) as total_paid
                FROM payments p
                JOIN students s ON p.student_id = s.student_id
                GROUP BY class_id, term
            ) p ON fs.class_id = p.class_id AND fs.term = p.term
        ''')
        outstanding = cursor.fetchone()[0] or 0

        # Payments today
        today = datetime.now().strftime('%Y-%m-%d')
        cursor.execute(db.convert_query('SELECT COUNT(*) FROM payments WHERE payment_date = ?'), (today,))
        payments_today = cursor.fetchone()[0]

        conn.close()
        return {
            'total_students': total_students,
            'total_collected': total_collected,
            'outstanding': outstanding,
            'payments_today': payments_today
        }

    @staticmethod
    def get_fees_per_class():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT c.name, SUM(p.amount) as total
            FROM payments p
            JOIN students s ON p.student_id = s.student_id
            JOIN classes c ON s.class_id = c.id
            GROUP BY c.id, c.name
            ORDER BY c.name
        ''')
        data = cursor.fetchall()
        conn.close()
        return data

    @staticmethod
    def get_students_owing():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT s.student_id, s.full_name, c.name as class_name,
                   SUM(fs.amount) - COALESCE(SUM(p.amount), 0) as balance
            FROM students s
            JOIN classes c ON s.class_id = c.id
            JOIN fees_structure fs ON fs.class_id = c.id
            LEFT JOIN payments p ON p.student_id = s.student_id AND p.term = fs.term
            GROUP BY s.student_id, s.full_name, c.name
            HAVING SUM(fs.amount) - COALESCE(SUM(p.amount), 0) > 0
            ORDER BY balance DESC
        ''')
        data = cursor.fetchall()
        conn.close()
        return data

    @staticmethod
    def get_fees_per_term():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT term, SUM(amount) as total
            FROM payments
            GROUP BY term
            ORDER BY term
        ''')
        data = cursor.fetchall()
        conn.close()
        return data

    @staticmethod
    def get_collection_trends():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        if db.use_postgres:
            cursor.execute('''
                SELECT TO_CHAR(payment_date::date, 'YYYY-MM') as month, SUM(amount) as total
                FROM payments
                WHERE payment_date >= (CURRENT_DATE - INTERVAL '12 months')
                GROUP BY month
                ORDER BY month
            ''')
        else:
            cursor.execute('''
                SELECT strftime('%Y-%m', payment_date) as month, SUM(amount) as total
                FROM payments
                WHERE payment_date >= date('now', '-12 months')
                GROUP BY strftime('%Y-%m', payment_date)
                ORDER BY month
            ''')
        data = cursor.fetchall()
        conn.close()
        
        labels = [row[0] for row in data]
        values = [row[1] for row in data]
        return {'labels': labels, 'data': values}

    @staticmethod
    def get_payment_methods_distribution():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT payment_method, COUNT(*) as count
            FROM payments
            GROUP BY payment_method
            ORDER BY count DESC
        ''')
        data = cursor.fetchall()
        conn.close()
        
        labels = [row[0] or 'Cash' for row in data]
        values = [row[1] for row in data]
        return {'labels': labels, 'data': values}

    @staticmethod
    def get_class_performance():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT c.name,
                   COALESCE(SUM(p.amount), 0) as collected,
                   SUM(fs.amount) - COALESCE(SUM(p.amount), 0) as outstanding
            FROM classes c
            JOIN fees_structure fs ON c.id = fs.class_id
            LEFT JOIN students s ON c.id = s.class_id
            LEFT JOIN payments p ON s.student_id = p.student_id AND p.term = fs.term
            GROUP BY c.id, c.name
            ORDER BY c.name
        ''')
        data = cursor.fetchall()
        conn.close()
        
        labels = [row[0] for row in data]
        collected = [row[1] for row in data]
        outstanding = [row[2] for row in data]
        return {'labels': labels, 'collected': collected, 'outstanding': outstanding}

    @staticmethod
    def get_outstanding_fees_by_class():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT c.name, SUM(fs.amount - COALESCE(p.total_paid, 0)) as outstanding
            FROM classes c
            JOIN fees_structure fs ON c.id = fs.class_id
            LEFT JOIN (
                SELECT s.class_id, p.term, SUM(p.amount) as total_paid
                FROM payments p
                JOIN students s ON p.student_id = s.student_id
                GROUP BY s.class_id, p.term
            ) p ON c.id = p.class_id AND fs.term = p.term
            GROUP BY c.id, c.name
            HAVING outstanding > 0
            ORDER BY outstanding DESC
        ''')
        data = cursor.fetchall()
        conn.close()
        
        labels = [row[0] for row in data]
        values = [row[1] for row in data]
        return {'labels': labels, 'data': values}

    @staticmethod
    def get_monthly_payment_patterns():
        db = Database()
        conn = db.get_connection()
        cursor = conn.cursor()
        
        if db.use_postgres:
            cursor.execute('''
                SELECT TO_CHAR(payment_date::date, 'YYYY-MM') as month, COUNT(*) as payments
                FROM payments
                WHERE payment_date >= (CURRENT_DATE - INTERVAL '12 months')
                GROUP BY month
                ORDER BY month
            ''')
        else:
            cursor.execute('''
                SELECT strftime('%Y-%m', payment_date) as month, COUNT(*) as payments
                FROM payments
                WHERE payment_date >= date('now', '-12 months')
                GROUP BY strftime('%Y-%m', payment_date)
                ORDER BY month
            ''')
        data = cursor.fetchall()
        conn.close()
        
        labels = [row[0] for row in data]
        values = [row[1] for row in data]
        return {'labels': labels, 'data': values}